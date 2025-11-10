import streamlit as st
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Διάσταση2 Aggregator (B,K,L only)", layout="wide")
st.title("📊 Sheet2→Sheet1: Aggregate by Διάσταση 2 (B) and write K & L")

uploaded = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"])

# ----- EXACT zero accounts you provided -----
ZERO_ACCOUNTS = {
    "50.00.00.0000","50.00.00.0001","50.00.00.0002","50.00.00.0003",
    "50.01.00.0000","50.01.01.0000","50.05.00.0000"
}

# Διάσταση 2 → Greek title (for matching Sheet1 rows by Τίτλος)
D2_TO_TITLE = {
    "--": "Προμηθευτές Capex πιστωτικά υπόλοιπα τέλους περιόδου",
    "01 - OpEx Payables": "Προμηθευτές Capex πιστωτικά υπόλοιπα τέλους περιόδου",
    "03 - Other Payables": "Προμηθευτές Capex πιστωτικά υπόλοιπα τέλους περιόδου",
    "100 - General B2B Invoices – Payments": "Προμηθευτές Capex πιστωτικά υπόλοιπα τέλους περιόδου",
    "110 - B2B Aging collections": "Προμηθευτές Capex πιστωτικά υπόλοιπα τέλους περιόδου",
    "2200 - Development Capex": "Προμηθευτές Capex πιστωτικά υπόλοιπα τέλους περιόδου",
    "300 - Financing Cashflows": "Προμηθευτές Capex πιστωτικά υπόλοιπα τέλους περιόδου",
    "02 - CapEx Payables": "Προμηθευτές πιστωτικά υπόλοιπα τέλους περιόδου",
    "04 - OpEx Advances": "Προμηθευτές χρεωστικά (προκαταβολές) υπόλοιπα τέλους περιόδου - Χρεώστες",
    "05 - CapEx Advances": "Προμηθευτές χρεωστικά (προκαταβολές) υπόλοιπα τέλους περιόδου - Προκαταβολές για αγορές Παγίων",
    "06 - Other Advances": "Προμηθευτές χρεωστικά (προκαταβολές) υπόλοιπα τέλους περιόδου - Χρεώστες"
}
# reverse: Greek title → Διάσταση 2 key
TITLE_TO_D2 = {v.strip(): k for k, v in D2_TO_TITLE.items()}

def find_col_exact(ws, name):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and str(v).strip() == name:
            return c
    return None

def find_col_contains(ws, needle):
    needle = needle.lower()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and needle in str(v).lower():
            return c
    return None

if uploaded:
    try:
        wb = load_workbook(uploaded)
        ws1 = wb.worksheets[0]   # Sheet1 (target)
        ws2 = wb.worksheets[1]   # Sheet2 (source)

        # ---- Delete duplicate column E if it's "Λογαριασμός λογιστικής"
        dupE = find_col_exact(ws1, "Λογαριασμός λογιστικής")
        if dupE: ws1.delete_cols(dupE)

        # ---- Sheet2 aggregation by positions: B, K, L
        B_col = 2   # Διάσταση 2
        K_col = 11  # column K
        L_col = 12  # column L

        aggK = {}  # d2 -> sum(K)
        aggL = {}  # d2 -> sum(L)
        for r in range(2, ws2.max_row + 1):
            d2 = ws2.cell(r, B_col).value
            if not d2: continue
            d2 = str(d2).strip()
            try: k_val = float(ws2.cell(r, K_col).value or 0)
            except: k_val = 0.0
            try: l_val = float(ws2.cell(r, L_col).value or 0)
            except: l_val = 0.0
            aggK[d2] = aggK.get(d2, 0.0) + k_val
            aggL[d2] = aggL.get(d2, 0.0) + l_val

        # ---- Sheet1: locate key columns
        title_col = find_col_exact(ws1, "Τίτλος")
        credit_col = find_col_exact(ws1, "Πιστωτικό Υπόλοιπο")  # this is K
        if credit_col is None: raise ValueError("Column 'Πιστωτικό Υπόλοιπο' not found in Sheet1.")
        debit_col = credit_col - 1                              # J = Χρεωστικό Υπόλοιπο
        L_pos = credit_col + 1                                  # L
        # Insert Διάσταση 2 (Source) after L
        ws1.insert_cols(L_pos + 1)
        src_col = L_pos + 1
        ws1.cell(1, src_col, "Διάσταση 2 (Source)")

        # Account code column (contains values like 50.00.00.0000)
        acct_col = find_col_contains(ws1, "Λογαριασμός")
        if acct_col is None:
            # fallback: try column B (common in your file)
            acct_col = 2

        # ---- Update rows
        for r in range(2, ws1.max_row + 1):
            acct = str(ws1.cell(r, acct_col).value or "").strip()

            # Zero ONLY if in your explicit list
            if acct in ZERO_ACCOUNTS:
                ws1.cell(r, debit_col, 0)   # J
                ws1.cell(r, credit_col, 0)  # K
                ws1.cell(r, L_pos,     0)   # L
                ws1.cell(r, src_col,   "")
                continue

            # Map by Greek title → Διάσταση 2 key
            d2_key = ""
            if title_col:
                title_val = str(ws1.cell(r, title_col).value or "").strip()
                d2_key = TITLE_TO_D2.get(title_val, "")

            # If we have a Διάσταση 2 key and aggregates, write them; else leave row untouched
            if d2_key and (d2_key in aggK or d2_key in aggL):
                if d2_key in aggK: ws1.cell(r, credit_col, aggK[d2_key])  # K
                if d2_key in aggL: ws1.cell(r, L_pos,     aggL[d2_key])   # L
                ws1.cell(r, src_col, d2_key)
            else:
                ws1.cell(r, src_col, "")

        # ---- Save
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        st.success("✅ Done. K & L updated from Sheet2 (B,K,L) and Διάσταση 2 (Source) filled. Zero-accounts set to 0.")
        st.download_button("⬇️ Download Updated Excel",
                           data=out,
                           file_name="Updated_" + uploaded.name,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"❌ Error: {e}")
