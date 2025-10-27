from fastapi import FastAPI, Request
from fastapi.responses import JSONResponse, PlainTextResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import base64
import io
import os
import pdfplumber
import openpyxl
from openai import OpenAI
from dotenv import load_dotenv

# ==========================
# LOAD ENVIRONMENT VARIABLES
# ==========================
load_dotenv()  # loads your .env with the API key
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# ==========================
# INIT APP
# ==========================
app = FastAPI(
    title="Vendor Reconciliation Analyzer",
    description="Receives base64 file uploads from Power Automate and classifies vendor type.",
    version="2.0"
)

# ==========================
# ENABLE CORS (optional but useful)
# ==========================
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================
# FIX FOR CHUNKED REQUESTS (Power Automate compatibility)
# ==========================
@app.middleware("http")
async def allow_chunked_requests(request: Request, call_next):
    """
    Power Automate sends chunked HTTP requests (without Content-Length),
    which FastAPI/Uvicorn normally rejects. This middleware reads the raw body
    and injects it back into the request to allow parsing.
    """
    if request.headers.get("transfer-encoding", "").lower() == "chunked":
        body = await request.body()
        request._body = body  # inject raw body back for Pydantic parsing
    response = await call_next(request)
    return response

# ==========================
# MODEL
# ==========================
class FilePayload(BaseModel):
    filename: str
    content: str  # base64-encoded file

# ==========================
# HELPER: Extract text from file
# ==========================
def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    filename = filename.lower()
    text = ""
    try:
        if filename.endswith(".pdf"):
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages[:5]:  # only first 5 pages
                    text += page.extract_text() or ""
        elif filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell])
                    text += row_text + "\n"
        else:
            text = file_bytes.decode(errors="ignore")
    except Exception as e:
        text = f"Error reading file: {e}"
    return text[:4000]  # keep within GPT input limits

# ==========================
# ROUTE
# ==========================
@app.post("/analyze")
async def analyze(payload: FilePayload):
    """
    Receives a file (base64 encoded), extracts readable text,
    and classifies it as ANDALUSIA / PORTO PETRO / ISHM / OTHER.
    """
    try:
        # Decode file content
        file_bytes = base64.b64decode(payload.content)
        text = extract_text_from_file(file_bytes, payload.filename)

        # Build prompt for classification
        prompt = f"""
        You are a classifier. Identify which of these appears in the text:
        ANDALUSIA, PORTO PETRO, IKOS SPANISH HOTEL MANAGEMENT, or OTHER.
        Return only one of those words. No explanations.

        Text:
        {text}
        """

        # GPT classification
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}]
        )

        result = response.choices[0].message.content.strip().upper()

        # Determine keyword
        if "ANDALUSIA" in result:
            keyword = "ANDALUSIA"
        elif "PORTO" in result:
            keyword = "PORTO PETRO"
        elif "SPANISH" in result or "IKOS SPANISH HOTEL" in result:
            keyword = "IKOS SPANISH HOTEL MANAGEMENT"
        else:
            keyword = "OTHER"

        return JSONResponse({"keyword": keyword})

    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

# ==========================
# ROOT ROUTE
# ==========================
@app.get("/")
def home():
    return PlainTextResponse("✅ FastAPI Analyzer is running and ready to receive POST /analyze requests.")

# ==========================
# ENTRY POINT (local run)
# ==========================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 8000)))
