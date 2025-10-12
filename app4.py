import io
import re
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from datetime import datetime

st.set_page_config(page_title="AP Overdue Email Manager", page_icon="💼", layout="wide")
st.title("🦖 AP-Rex — The Invoice Hunter")

# ================= HELPER FUNCTIONS =================
def safe_excel_to_df(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = next((wb[name] for name in wb.sheetnames if wb[name].max_row > 1 and wb[name].max_column > 1), wb.active)
    data = []
    for row in ws.values:
        safe_row = ["" if cell is None else str(cell) for cell in row]
        data.append(safe_row)
    if not data:
        raise ValueError("Excel file is empty.")
    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(data[0])]
    seen = {}
    unique_headers = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)
    return pd.DataFrame(data[1:], columns=unique_headers)

def combine_emails(df):
    """Auto-detect and combine all email-like columns into one unified column."""
    email_cols = []
    for c in df.columns:
        c_low = c.lower()
        if any(k in c_low for k in ["email", "e-mail", "correo", "ηλεκτρον", "διευθυν"]):
            email_cols.append(c)
    if not email_cols:
        st.warning("⚠️ No email columns found in file.")
        return df

    df["combined_emails"] = df[email_cols].apply(
        lambda r: "; ".join(
            sorted({str(x).strip() for x in r if isinstance(x, str) and "@" in x and len(x.strip()) > 5})
        ),
        axis=1
    )
    return df

# ---- FIX: safe flattener for email series (avoids .sum() on lists) ----
def _collect_emails(series: pd.Series) -> str:
    """Given a Series with 'a; b; c' strings, return a '; ' joined unique cleaned list."""
    if series is None or series.empty:
        return ""
    s = series.dropna().astype(str).str.split(";")
    # flatten safely
    flat = s.explode().dropna().astype(str).str.strip()
    # keep valid-looking emails only
    flat = flat[(flat.str.contains("@")) & (flat.str.len() > 5)]
    return "; ".join(sorted(flat.unique()))

# ================= MAIN APP =================
uploaded = st.file_uploader("📦 Upload Excel (.xlsx)", type=["xlsx"])

if uploaded:
    try:
        df = safe_excel_to_df(uploaded)

        # --- FILTER BASE EXCEL ---
        if "document" in df.columns:
            df = df[~df["document"].str.contains("F&B", case=False, na=False)]
        if "type" in df.columns:
            df = df[df["type"].str.upper() == "XPI"]
        for c in [c for c in df.columns if "payment_method" in c]:
            df = df[~df[c].str.lower().isin(["downpayment", "direct debit", "cash", "credit card"])]

        agreed_col = next((c for c in ["agreed", "agreeded"] if c in df.columns), None)
        if agreed_col:
            df[agreed_col] = pd.to_numeric(df[agreed_col], errors="coerce").fillna(0)
            df = df[df[agreed_col] == 0]

        st.session_state.df_session = df
        st.success(f"✅ Excel loaded and filtered: {len(df)} rows")
        st.dataframe(df.head(20), use_container_width=True)

        prompt = st.text_area("Type your request (supports multi-line):")
        df = st.session_state.df_session.copy()

        # ========== 1️⃣ SHOW OVERDUE INVOICES ==========
        if prompt.lower().startswith("show overdue invoices"):
            m = re.search(r"as of\s+(\d{4}-\d{2}-\d{2})", prompt)
            ref_date = pd.to_datetime(m.group(1)) if m else datetime.today()

            if "due_date" not in df.columns:
                st.error("⚠️ No due_date column found in Excel.")
            else:
                df["due_date_parsed"] = pd.to_datetime(df["due_date"], errors="coerce")
                overdue_df = df[df["due_date_parsed"] < ref_date].copy()

                vendor_col = next((c for c in ["vendor_name", "supp_name", "supplier", "vendor"] if c in overdue_df.columns), None)
                if "open_amount" in overdue_df.columns:
                    overdue_df["open_amount"] = pd.to_numeric(overdue_df["open_amount"], errors="coerce").fillna(0)
                elif "open_amount_in_base_cur" in overdue_df.columns:
                    overdue_df["open_amount"] = pd.to_numeric(overdue_df["open_amount_in_base_cur"], errors="coerce").fillna(0)
                else:
                    overdue_df["open_amount"] = 0

                st.session_state.filtered_df = overdue_df
                total_overdue = overdue_df["open_amount"].sum()

                st.warning(f"⚠️ Found {len(overdue_df)} overdue invoices as of {ref_date.date()}")
                st.write(f"💰 **Total overdue amount:** {total_overdue:,.2f} EUR")
                display_cols = [c for c in [vendor_col, "document", "due_date", "open_amount"] if c]
                st.dataframe(overdue_df[display_cols], use_container_width=True)

        # ========== 2️⃣ GET EMAILS FOR CURRENT FILTER ==========
        elif "get emails for current filter" in prompt.lower():
            if "filtered_df" not in st.session_state or st.session_state.filtered_df.empty:
                st.error("⚠️ No active filter. Run 'show overdue invoices ...' first.")
            else:
                df = combine_emails(st.session_state.filtered_df.copy())
                if "country" not in df.columns:
                    df["country"] = "other"
                df["lang"] = df["country"].str.lower().apply(
                    lambda x: "ES" if any(k in x for k in ["spain", "es", "esp", "españa"]) else "EN"
                )

                if "combined_emails" not in df.columns:
                    st.error("⚠️ No email columns found to combine.")
                else:
                    # ---- FIX: use _collect_emails instead of .str.split().sum() ----
                    es_emails = _collect_emails(df.loc[df["lang"] == "ES", "combined_emails"])
                    en_emails = _collect_emails(df.loc[df["lang"] == "EN", "combined_emails"])

                    st.write(f"📅 Filtered overdue invoices: {len(df)} rows")
                    st.write("🇪🇸 **Spanish vendor emails (copy for Outlook)**")
                    st.code(es_emails or "No Spanish emails found", language="text")
                    st.write("🇬🇧 **English vendor emails (copy for Outlook)**")
                    st.code(en_emails or "No English emails found", language="text")

        # ========== 3️⃣ GET ALL EMAILS (GLOBAL) ==========
        elif "give me all spanish and english emails" in prompt.lower():
            df = combine_emails(st.session_state.df_session.copy())
            if "country" not in df.columns:
                df["country"] = "other"

            df["lang"] = df["country"].str.lower().apply(
                lambda x: "ES" if any(k in x for k in ["spain", "es", "esp", "españa"]) else "EN"
            )

            if "combined_emails" not in df.columns:
                st.error("⚠️ No email columns found to combine.")
            else:
                # ---- FIX: use _collect_emails instead of .str.split().sum() ----
                es_emails = _collect_emails(df.loc[df["lang"] == "ES", "combined_emails"])
                en_emails = _collect_emails(df.loc[df["lang"] == "EN", "combined_emails"])

                st.write("🇪🇸 **All Spanish vendor emails (copy for Outlook)**")
                st.code(es_emails or "No Spanish emails found", language="text")
                st.write("🇬🇧 **All English vendor emails (copy for Outlook)**")
                st.code(en_emails or "No English emails found", language="text")

        # ========== 4️⃣ GET OPEN AMOUNTS EMAILS BY LANGUAGE ==========
        elif "give me the open amounts emails" in prompt.lower():
            df = combine_emails(st.session_state.df_session.copy())
            if "country" not in df.columns:
                df["country"] = "other"

            df["lang"] = df["country"].str.lower().apply(
                lambda x: "ES" if any(k in x for k in ["spain", "es", "esp", "españa"]) else "EN"
            )

            if "combined_emails" not in df.columns:
                st.error("⚠️ No email columns found to combine.")
            else:
                # ---- FIX: use _collect_emails instead of .str.split().sum() ----
                es_emails = _collect_emails(df.loc[df["lang"] == "ES", "combined_emails"])
                en_emails = _collect_emails(df.loc[df["lang"] == "EN", "combined_emails"])

                st.write("🇪🇸 **Open amounts — Spanish vendor emails**")
                st.code(es_emails or "No Spanish emails found", language="text")
                st.write("🇬🇧 **Open amounts — English vendor emails**")
                st.code(en_emails or "No English emails found", language="text")

        # ========== 5️⃣ FIND INVALID OR MISSING EMAILS ==========
        elif "find invalid or missing emails" in prompt.lower():
            email_cols = [c for c in df.columns if any(k in c.lower() for k in ["email", "e-mail", "correo", "ηλεκτρον", "διευθυν"])]
            if not email_cols:
                st.error("⚠️ No email columns found.")
            else:
                invalid_df = df[df[email_cols].apply(lambda x: all(not re.search(r"@.+\.", str(i)) for i in x), axis=1)]
                vendor_col = next((c for c in ["vendor_name", "supp_name", "supplier", "vendor"] if c in df.columns), None)
                cols = [vendor_col] + email_cols if vendor_col else email_cols
                st.dataframe(invalid_df[cols].drop_duplicates(), use_container_width=True)

        # ========== 6️⃣ ADD MULTIPLE EMAILS ==========
        elif prompt.lower().startswith("add multiple emails:"):
            lines = prompt.split("\n")[1:]
            vendor_col = next((c for c in ["vendor_name", "supp_name", "supplier", "vendor"] if c in df.columns), None)
            if not vendor_col:
                st.error("⚠️ No vendor column found.")
            else:
                updates = []
                for line in lines:
                    if ":" in line:
                        name, emails = line.split(":", 1)
                        name, emails = name.strip(), emails.strip()
                        mask = df[vendor_col].str.lower().eq(name.lower())
                        if mask.any():
                            for c in [c for c in df.columns if "email" in c.lower()]:
                                df.loc[mask, c] = emails
                            updates.append(name)
                st.session_state.df_session = df
                st.success(f"✅ Updated emails for: {', '.join(updates)}")

        # ========== 7️⃣ SHOW TOTAL OPEN AMOUNTS ==========
        elif "show total open amounts" in prompt.lower():
            amount_col = "open_amount" if "open_amount" in df.columns else "open_amount_in_base_cur"
            total = pd.to_numeric(df[amount_col], errors="coerce").sum()
            st.write(f"💰 Total open amount: {total:,.2f} EUR")

        else:
            if prompt.strip():
                st.warning("⚠️ Unknown command. Try one of the email or invoice prompts above.")

    except Exception as e:
        st.error(f"❌ Error: {e}")
